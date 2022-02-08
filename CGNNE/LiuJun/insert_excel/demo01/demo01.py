"""
1、读取已有的excel文档
2、根据时间的不同，对不同的sheet进行插入数据
3、对输入的数据进行判断，如果不行，则不让输入，如果可以，则输入成功
"""
from tkinter import *
from tkinter import ttk, messagebox
from tkinter.filedialog import askopenfilename

import xlrd
from openpyxl import load_workbook

class StationInfo(object):
    def __init__(self,id,station_name,station_capacity,station_generator_num):
        self.id=id
        self.station_name=station_name
        self.station_capacity=station_capacity
        self.station_generator_num=station_generator_num
class StationInsertInfo(object):
    def __init__(self,stationName,station_generator_num):
        self.id=id
        self.stationName=stationName
        self.station_generator_num=station_generator_num

class Fault_Input_Frame(Frame):
    def __init__(self,root1=None,station_list=None,filepath=None):
        super().__init__(root1)
        self.station_list=station_list
        self.stationGenratorNum=0
        self.filepath=filepath
        station_name_list=[]
        for i in self.station_list:
            station_name_list.append(i.station_name)
        stationBox = ttk.Combobox(self,  values=station_name_list,width=20)
        stationBox.grid(row=0, column=1,pady=10)
        L1 = Label(self, text="请输入故障机组号:")
        L1.grid(row=1, column=0, padx=20)
        E1 = Entry(self, width=20)
        E1.grid(row=1, column=1)
        B3= Button(self, text="插入数据", command=lambda: self.getStationGenratorNum(), height=1)
        B3.grid(row=2, column=1, padx=10)

    def getStationGenratorNum(self):
        for i in self.station_list:
            if i.station_name==self.children["!combobox"].get():
                self.stationGenratorNum=i.station_generator_num

        num=self.children["!entry"].get()
        try:
            num=int(num)
        except Exception:
            messagebox.showinfo('提示', '输入有误')
        if num>self.stationGenratorNum:
            messagebox.showinfo('提示', '输入有误')
        else:
            workbook=xlrd.open_workbook(self.filepath)
            sheet = workbook.sheet_by_name("sheet")
            row_num = sheet.nrows
            id=row_num
            stationName=self.children["!combobox"].get()
            station_generator_num=num
            wb=load_workbook(filename=self.filepath)
            ws=wb["sheet"]
            ws.cell(row=row_num+1,column=1,value=id)
            ws.cell(row=row_num+1,column=2,value=stationName)
            ws.cell(row=row_num+1,column=3,value=station_generator_num)

            wb.save(self.filepath)
            messagebox.showinfo('提示', '输入成功')




class input_file_frame(Frame):
    def __init__(self,root=None):
        super().__init__(root)
        self.path1 = StringVar()
        self.path2 = StringVar()
        self.filepath=None
        L1 = Label(self, text="读取故障记录表格：")
        L1.grid(row=0, column=0,padx=20)
        E1 = Entry(self, textvariable=self.path1, width=30)
        E1.grid(row=0, column=1)
        # command 后面的函数不能有参数，否则会直接执行，跳出文本选择框，需要在函数前加lambda，变成匿名函数
        B1 = Button(self, text="...", command=lambda: self.selectPath01(self.path1), height=1)
        B1.grid(row=0, column=2, padx=10)
        B2 = Button(self, text="故障录入", command=lambda: self.input(), height=1)
        B2.grid(row=1, column=1, padx=1)

    def selectPath01(self,path1):
        path_ = askopenfilename()
        path1.set(path_)
        self.filepath01=path_
    def input(self):
        filepath01=self.filepath01
        work_book02 = xlrd.open_workbook(self.filepath01)
        sheet02 = work_book02.sheet_by_name("风场基本数据")
        row_num02 = sheet02.nrows
        station_list = []
        for i in range(1, row_num02):
            id, station_name, station_capacity, station_generator_num = sheet02.row_values(i, 0, 4)
            station_capacity = int(station_capacity)
            station_generator_num = int(station_generator_num)
            s = StationInfo(id, station_name, station_capacity, station_generator_num)
            station_list.append(s)
        root1=Tk(className="数据输入")
        root1.geometry("500x100")
        fault_input_frame=Fault_Input_Frame(root1,station_list,filepath01)
        fault_input_frame.pack()

root=Tk(className="故障记录输入")
root.geometry("500x100")

if_frame=input_file_frame(root)
if_frame.pack()

root.mainloop()