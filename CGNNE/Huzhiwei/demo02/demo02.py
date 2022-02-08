
"""
1、读取excel表格
2、生成相应对象
3、数据分析
"""
from tkinter import *
from tkinter.filedialog import askopenfilename
import matplotlib
import matplotlib.pyplot as plt
import openpyxl

import xlrd

class Fault(object):
    def __init__(self,id,company,station,generator_num,generator_no,
                 generator_factory,generator_model,fault_name,info,
                 fault_code,fault_description,fault_kind,
                 fault_start_time,maintanence_start_time,recover_time,
                 fault_first_location,fault_second_location,fault_third_location,
                 check_list,fault_reason,maintantence_level,maintanence_kind,
                 work_ticket_num,maintanence_movement,maintanence_object,
                 object_name,object_kind,object_num,object_danwei,object_factory,
                 test_method,fault_analysis_report_num,lost_power,
                 maintanence_cost,total,people_name,people_num):
        self.id=id
        self.company=company
        self.station=station
        self.generator_num=generator_num
        self.generator_factory=generator_factory
        self.generator_no=generator_no
        self.generator_model=generator_model
        self.fault_name=fault_name
        self.info=info
        self.fault_code=fault_code
        self.fault_description=fault_description
        self.fault_kind=fault_kind
        self.fault_start_time=fault_start_time
        self.maintanence_start_time=maintanence_start_time
        self.recover_time=recover_time
        self.fault_first_location=fault_first_location
        self.fault_second_location=fault_second_location
        self.fault_third_location=fault_third_location
        self.check_list=check_list
        self.fault_reason=fault_reason
        self.maintanence_level=maintantence_level
        self.maintanence_kind=maintanence_kind
        self.work_ticket_num=work_ticket_num
        self.maintanence_movement=maintanence_movement
        self.maintanence_object=maintanence_object
        self.object_name=object_name
        self.object_kind=object_kind
        self.object_num=object_num
        self.object_danwei=object_danwei
        self.object_factory=object_factory
        self.test_method=test_method
        self.fault_analysis_report_num=fault_analysis_report_num
        self.lost_power=lost_power
        self.maintanence_cost=maintanence_cost
        self.total=self.get_total(total)
        self.people_name=people_name
        self.people_num=people_num

    def get_total(self, total):
        if total=="":
            num=0.0
        elif total=="0":
            num=0.0
        else:
            num=float(total)
        return num
class StationInfo(object):
    def __init__(self,id,station_name,station_capacity,station_generator_num):
        self.id=id
        self.station_name=station_name
        self.station_capacity=station_capacity
        self.station_generator_num=station_generator_num
        self.total_fault_count=0
        self.total_fault_time=0.0
        self.total_lost=0.0
#获取每个对象的时间差
def getTime(start_time,end_time):
    start_time=xlrd.xldate.xldate_as_datetime(start_time, 0)
    end_time=xlrd.xldate.xldate_as_datetime(end_time, 0)
    time=end_time-start_time
    return time
# 风机厂信息
class Factory_Info(object):
    def __init__(self, id, name,num,capacity):
        self.id = id
        self.name = name
        self.num=num
        self.capcity=capacity
        self.total_fault_count=0
        self.total_falut_time=0.0
        self.total_maintanece_time=0.0
        self.power_lost=0.0
        self.total_cost=0.0
#获取故障总次数，故障修复时间，经济损失
class Fault_info(object):
    def __init__(self,id,fault_first_location):
        self.id=id
        self.fault_first_location=fault_first_location
        self.total_fault_count=0
        self.total_fault_time=0.0
        self.maintanence_time=0.0
        self.total_lost=0.0
class Maintanence_info(object):
    def __init__(self,id,maintanence_level):
        self.id=id
        self.maintanence_level=maintanence_level
        self.total_fault_count = 0
        self.total_fault_time = 0.0
        self.maintanence_time = 0.0
        self.total_lost = 0.0
def plot(title, x_label=None, y_value=None, y_label=None, hasaverage=None):
    """

    :param title: 图纸名称
    :param x_label: x的坐标值
    :param y_value: 与x对应的y的值
    :param y_label: y的坐标值
    :param hasaverage: 是否画平均线
    :return:
    """

    if y_label is None:
        y_label = []
    if y_value is None:
        y_value = []
    if x_label is None:
        x_label = []
    # 横坐标显示中文
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
    # font = {'family': 'Microsoft YaHei', 'size': '14'}
    # matplotlib.rc("font", family="Microsoft YaHei",  size="14")

    chang=len(x_label)*1.2
    #指定画布大小
    plt.figure(figsize=(chang, 4))


    #设置x轴坐标整体名称
    plt.xlabel("")
    # 添加Y轴标签
    plt.ylabel("")
    #设置柱形图的x坐标值，和对应的y值
    #设置柱子的宽度width
    #设置柱子的颜色color
    color=[]
    for i in range(len(x_label)):
        if i<3:
            color.append('r')
        elif i>(len(x_label)-4):
            color.append('g')
        else:
            color.append('b')
    plt.bar(x_label, y_value, width=0.3,color=color)
    if title=="TBA":
        # 设置图表名称
        plt.title(title, loc="center", pad=10)
        plt.ylim([0.96,1])
        for a, b, i in zip(x_label, y_value, range(len(x_label))):
            plt.text(a, b, "%.2f%%" % (y_value[i] * 100), ha='center', fontsize=10)
    else:
    # 通过zip 函数函数，遍历x，y列表
    #在柱形图上，设置文字
    # plt.text 函数
    # 设置图表名称
        plt.title(title, loc="center")
        for a, b, i in zip(x_label, y_value, range(len(x_label))):
            plt.text(a, b, "%.1f" % y_value[i], ha='center', fontsize=10)
    #画平均值线
    y_sum=0
    for i in y_value:
        y_sum+=i
    y_average=y_sum/len(y_value)
    # 添加水平直线
    plt.axhline(y=y_average, ls="-", c="orange",label="平均值")
    # 打开坐标网格
    plt.grid(axis="y",ls="-", c='grey', )
    #图例设置
    plt.legend(loc='lower center')
    plt.show()
class Draw_Picture_Frame01(Frame):
    def __init__(self,root=None,station_list=None):
        super().__init__(root)
        self.station_list=station_list
        B1 = Button(self, text="MTBF", command=lambda: self.draw_picture("MTBF"), height=1)
        B1.grid(row=0, column=1, padx=10, pady=30)
        B2 = Button(self, text="MTTR", command=lambda: self.draw_picture("MTTR"), height=1)
        B2.grid(row=1, column=1, padx=10, pady=30)
        B3 = Button(self, text="TBA", command=lambda: self.draw_picture("TBA"), height=1)
        B3.grid(row=2, column=1, padx=10, pady=30)
        B4 = Button(self, text="单位容量故障次数", command=lambda: self.draw_picture("单位容量故障次数"), height=1)
        B4.grid(row=3, column=1, padx=10, pady=30)
        B5 = Button(self, text="单位容量经济损失", command=lambda: self.draw_picture("单位容量经济损失"), height=1)
        B5.grid(row=4, column=1, padx=10, pady=30)

    def draw_picture(self, param):
        x = list()
        y = list()
        if param == "MTBF":
            self.station_list.sort(key=lambda x: x.MTBF)
            for i in self.station_list:
                x.append(i.station_name.strip("风电场"))
                y.append(i.MTBF)
        elif param == "MTTR":
            self.station_list.sort(key=lambda x: x.MTTR)
            for i in self.station_list:
                x.append(i.station_name.strip("风电场"))
                y.append(i.MTTR)
        elif param == "TBA":
            self.station_list.sort(key=lambda x: x.TBA)
            for i in self.station_list:
                x.append(i.station_name.strip("风电场"))
                y.append(i.TBA)
        elif param == "单位容量故障次数":
            self.station_list.sort(key=lambda x: x.fault_count_per_capacity)
            for i in self.station_list:
                x.append(i.station_name.strip("风电场"))
                y.append(i.fault_count_per_capacity)
        else:
            self.station_list.sort(key=lambda x: x.lost_per_capacity)
            for i in self.station_list:
                x.append(i.station_name.strip("风电场"))
                y.append(i.lost_per_capacity)

        plot(param, x, y)

class Draw_Picture_Frame02(Frame):
    def __init__(self,root=None,factory_list=None):
        super().__init__(root)
        self.factory_list=factory_list
        B1 = Button(self, text="MTBF", command=lambda: self.draw_picture("MTBF"), height=1)
        B1.grid(row=0, column=1, padx=10, pady=30)
        B2 = Button(self, text="MTTR", command=lambda: self.draw_picture("MTTR"), height=1)
        B2.grid(row=1, column=1, padx=10, pady=30)
        B3 = Button(self, text="TBA", command=lambda: self.draw_picture("TBA"), height=1)
        B3.grid(row=2, column=1, padx=10, pady=30)
        B4 = Button(self, text="单位容量故障次数", command=lambda: self.draw_picture("单位容量故障次数"), height=1)
        B4.grid(row=3, column=1, padx=10, pady=30)
        B5 = Button(self, text="单位容量经济损失", command=lambda: self.draw_picture("单位容量经济损失"), height=1)
        B5.grid(row=4, column=1, padx=10, pady=30)

    def draw_picture(self, param):
        x = list()
        y = list()
        if param == "MTBF":
            self.factory_list.sort(key=lambda x: x.MTBF)
            for i in self.factory_list:
                x.append(i.name)
                y.append(i.MTBF)
        elif param == "MTTR":
            self.factory_list.sort(key=lambda x: x.MTTR)
            for i in self.factory_list:
                x.append(i.name)
                y.append(i.MTTR)
        elif param == "TBA":
            self.factory_list.sort(key=lambda x: x.TBA)
            for i in self.factory_list:
                x.append(i.name)
                y.append(i.TBA)
        elif param == "单位容量故障次数":
            self.factory_list.sort(key=lambda x: x.fault_count_per_capacity)
            for i in self.factory_list:
                x.append(i.name)
                y.append(i.fault_count_per_capacity)
        else:
            self.factory_list.sort(key=lambda x: x.lost_per_capacity)
            for i in self.factory_list:
                x.append(i.name)
                y.append(i.lost_per_capacity)

        plot(param,x,y)
class Make_Excel_Frame(Frame):
    def __init__(self, root=None, fault_info_list=None, maintanence_level_list=None):
        super().__init__(root)
        self.fault_info_list=fault_info_list
        self.maintanence_level_list=maintanence_level_list
        B1 = Button(self, text="按故障分布统计制表", command=lambda: self.make_excel("fault"), height=1)
        B1.grid(row=0, column=1, padx=10, pady=30)
        B2 = Button(self, text="按维护等级制表", command=lambda: self.make_excel("maintanence"), height=1)
        B2.grid(row=1, column=1, padx=10, pady=30)

    def make_excel(self, param):
        wb = openpyxl.Workbook()
        ws = wb.active  # 默认插在最后
        ws.title = 'sheet1'
        if param=="fault":
            ws['A1']="故障位置一级"
            ws['B1']="故障次数"
            ws['C1']="故障停机小时"
            ws['D1']="维修时间   "
            ws['E1']="故障损失合计"
            row_num=2
            for i in self.fault_info_list:
                ws.cell(row_num,1).value=i.fault_first_location
                ws.cell(row_num,2).value=i.total_fault_count
                ws.cell(row_num,3).value=round(i.total_fault_time/3600,2)
                ws.cell(row_num,4).value=round(i.maintanence_time/3600,2)
                ws.cell(row_num,5).value=i.total_lost
                row_num+=1
            wb.save('故障统计.xlsx')
        else:
            ws['A1'] = "维修等级"
            ws['B1'] = "故障次数"
            ws['C1'] = "故障停机小时"
            ws['D1'] = "维修时间"
            ws['E1'] = "故障损失合计"
            row_num = 2
            for i in self.maintanence_level_list:
                ws.cell(row_num, 1).value = i.maintanence_level
                ws.cell(row_num, 2).value = i.total_fault_count
                ws.cell(row_num, 3).value = round(i.total_fault_time/3600,2)
                ws.cell(row_num, 4).value = round(i.maintanence_time/3600,2)
                ws.cell(row_num, 5).value = i.total_lost
                row_num += 1
            wb.save('维修统计.xlsx')
class Select_Frame(Frame):
    def __init__(self,root=None,fault_list=None,station_list=None,factory_list=None,fault_info_list=None,maintanence_info_list=None):
        super().__init__(root)
        self.fault_list=fault_list
        self.station_list=station_list
        self.factory_list=factory_list
        self.fault_info_list=fault_info_list
        self.maintanence_info_list=maintanence_info_list
        B1 = Button(self, text="按场站统计", command=lambda: self.station_draw(), height=1)
        B1.grid(row=0, column=0, padx=10, pady=30)
        B2= Button(self, text="按风机厂家统计", command=lambda: self.factory_draw(), height=1)
        B2.grid(row=1, column=0, padx=10, pady=30)
        B3 = Button(self, text="整体故障统计", command=lambda: self.fault_draw(), height=1)
        B3.grid(row=2, column=0, padx=10, pady=30)
        

    def station_draw(self):
        root2 = Tk(className="画图")
        root2.geometry("200x500")
        draw_picture_frame = Draw_Picture_Frame01(root2,self.station_list)
        draw_picture_frame.pack()

    def factory_draw(self):
        root2 = Tk(className="画图")
        root2.geometry("200x500")
        draw_picture_frame = Draw_Picture_Frame02(root2, self.factory_list)
        draw_picture_frame.pack()

    def fault_draw(self):
        root3 = Tk(className="制表")
        root3.geometry("200x200")
        make_excel_frame = Make_Excel_Frame(root3, self.fault_info_list,self.maintanence_info_list)
        make_excel_frame.pack()


class input_file_frame(Frame):
    def __init__(self,root=None):
        super().__init__(root)
        self.path1 = StringVar()
        self.path2 = StringVar()
        self.filepath=None
        L1 = Label(self, text="输入故障数据文件:")
        L1.grid(row=0, column=0,pady=30,padx=20)
        E1 = Entry(self, textvariable=self.path1, width=30)
        E1.grid(row=0, column=1,pady=30)
        # command 后面的函数不能有参数，否则会直接执行，跳出文本选择框，需要在函数前加lambda，变成匿名函数
        B1 = Button(self, text="...", command=lambda: self.selectPath01(self.path1), height=1)
        B1.grid(row=0, column=2, padx=10,pady=30)
        L2 = Label(self, text="输入场站基本文件:")
        L2.grid(row=1, column=0, pady=30, padx=20)
        E2 = Entry(self, textvariable=self.path2, width=30)
        E2.grid(row=1, column=1, pady=30)
        # command 后面的函数不能有参数，否则会直接执行，跳出文本选择框，需要在函数前加lambda，变成匿名函数
        B1 = Button(self, text="...", command=lambda: self.selectPath02(self.path2), height=1)
        B1.grid(row=1, column=2, padx=10, pady=30)
        B2=Button(self,text="分析数据",command=lambda :self.read_excel(),height=1)
        B2.grid(row=2,column=1,padx=10)


    def selectPath01(self,path1):
        path_ = askopenfilename()
        path1.set(path_)
        self.filepath01=path_
    def selectPath02(self,path2):
        path_ = askopenfilename()
        path2.set(path_)
        self.filepath02=path_

    def read_excel(self):
        work_book = xlrd.open_workbook(self.filepath01)
        sheet = work_book.sheet_by_name("故障填报主表")
        row_num = sheet.nrows
        fault_list = []
        for i in range(4, row_num):
            id, company, station, generator_num, generator_no, generator_factory, generator_model, fault_name, info, fault_code, fault_description, fault_kind, fault_start_time, maintanence_start_time, recover_time, fault_first_location, fault_second_location, fault_third_location, check_list, fault_reason, maintantence_level, maintanence_kind, work_ticket_num, maintanence_movement, maintanence_object, object_name, object_kind, object_num, object_danwei, object_factory, test_method, fault_analysis_report_num, lost_power, maintanence_cost, total, people_name, people_num = sheet.row_values(
                i, 0, 37)
            f = Fault(id, company, station, generator_num, generator_no, generator_factory, generator_model, fault_name,
                      info, fault_code, fault_description, fault_kind, fault_start_time, maintanence_start_time,
                      recover_time, fault_first_location, fault_second_location, fault_third_location, check_list,
                      fault_reason, maintantence_level, maintanence_kind, work_ticket_num, maintanence_movement,
                      maintanence_object, object_name, object_kind, object_num, object_danwei, object_factory,
                      test_method, fault_analysis_report_num, lost_power, maintanence_cost, total, people_name,
                      people_num)
            fault_list.append(f)
        work_book02 = xlrd.open_workbook(self.filepath02)
        sheet02 = work_book02.sheet_by_name("Sheet1")
        row_num02 = sheet02.nrows
        station_list = []
        for i in range(1, row_num02):
            id, station_name, station_capacity, station_generator_num = sheet02.row_values(i, 0, 4)
            station_capacity = int(station_capacity)
            station_generator_num = int(station_generator_num)
            s = StationInfo(id, station_name, station_capacity, station_generator_num)
            station_list.append(s)
        sheet02_02 = work_book02.sheet_by_name("Sheet2")
        row_num02_02 = sheet02_02.nrows
        factory_list = []
        for i in range(1, row_num02_02):
            id, name, num, capacity = sheet02_02.row_values(i)
            id = int(id)
            num = int(num)
            f = Factory_Info(id, name, num, capacity)
            factory_list.append(f)
        fault_info_set = set()
        maintantence_info_set = set()
        # 获取故障总次数，故障修复时间，经济损失
        for i in fault_list:
            fault_info_set.add(i.fault_first_location)
            maintantence_info_set.add(i.maintanence_level)
            if (i.info == "非计划停机"):
                for j in station_list:
                    if i.station == j.station_name:
                        j.total_fault_count += 1
                        j.total_fault_time += getTime(i.fault_start_time, i.recover_time).total_seconds()
                        j.total_lost += i.total
                for j in factory_list:
                    if i.generator_factory == j.name:
                        j.total_fault_count += 1
                        j.total_falut_time += getTime(i.fault_start_time, i.recover_time).total_seconds()
                        j.total_maintanece_time += getTime(i.maintanence_start_time, i.recover_time).total_seconds()
                        j.total_cost += i.total
                        if i.lost_power == "":
                            pass
                        else:
                            j.power_lost += float(i.lost_power)
        id = 1
        fault_info_list = []
        for i in fault_info_set:
            f = Fault_info(id, i)
            fault_info_list.append(f)
        id = 1
        maintantence_info_list = []
        for i in maintantence_info_set:
            m = Maintanence_info(id, i)
            maintantence_info_list.append(m)
        # 开始计算MTBF，MTTR，TBA，单位容量故障次数，单位容量经济损失
        for i in station_list:
            if i.total_fault_count == 0:
                i.MTBF = round(30 * 24 * i.station_generator_num, 2)
                i.MTTR = round(i.total_fault_time / 3600, 2)
            else:
                i.MTBF = round(30 * 24 * i.station_generator_num / i.total_fault_count, 2)
                i.MTTR = round((i.total_fault_time / 3600) / i.total_fault_count, 2)
            i.TBA = round(
                (30 * 24 * i.station_generator_num - i.total_fault_time / 3600) / (30 * 24 * i.station_generator_num),
                4)
            i.fault_count_per_capacity = i.total_fault_count / i.station_capacity
            i.lost_per_capacity = i.total_lost / i.station_capacity
        for i in factory_list:
            if i.total_fault_count == 0:
                i.MTBF = round(30 * 24 * i.num, 1)
                i.MTTR = round(i.total_fault_count / 3600, 1)
            else:
                i.MTBF = round(30 * 24 * i.num / i.total_fault_count, 1)
                i.MTTR = round((i.total_falut_time / 3600) / i.total_fault_count, 1)
            i.TBA = round((30 * 24 * i.num - i.total_falut_time / 3600) / (30 * 24 * i.num), 4)
            i.fault_count_per_capacity = round(i.total_fault_count / i.capcity, 1)
            i.lost_per_capacity = round(i.total_cost / i.capcity, 2)

        # 开始计算故障总次数和维修等级次数
        for i in fault_list:
            for j in fault_info_list:
                if i.fault_first_location == j.fault_first_location:
                    j.total_fault_count += 1
                    if i.fault_start_time == "":
                        pass
                    else:
                        j.total_fault_time += getTime(i.fault_start_time, i.recover_time).total_seconds()
                    if i.fault_start_time == "":
                        pass
                    else:
                        j.maintanence_time += getTime(i.maintanence_start_time, i.recover_time).total_seconds()
                    j.total_lost += i.total
        for i in fault_list:
            for j in maintantence_info_list:
                if i.maintanence_level == j.maintanence_level:
                    j.total_fault_count += 1
                    if i.fault_start_time == "":
                        pass
                    else:
                        j.total_fault_time += getTime(i.fault_start_time, i.recover_time).total_seconds()
                    if i.fault_start_time == "":
                        pass
                    else:
                        j.maintanence_time += getTime(i.maintanence_start_time, i.recover_time).total_seconds()
                    j.total_lost += i.total
        root1=Tk(className="选择画图依据")
        root1.geometry("200x300")
        select_frame=Select_Frame(root1,fault_list,station_list,factory_list,fault_info_list,maintantence_info_list)
        select_frame.pack()
root=Tk(className="Excel文件处理")
root.geometry("500x300")
if_frame=input_file_frame(root)
if_frame.pack()

root.mainloop()